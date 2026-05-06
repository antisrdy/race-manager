"""Tests for utility functions."""
import io
import csv
import pytest
import openpyxl
from datetime import datetime
from app import compute_elapsed, normalize_runner_row, parse_upload


class TestComputeElapsed:
    """Tests for the compute_elapsed function."""
    
    def test_compute_elapsed_valid_times(self):
        """Test computing elapsed time with valid start and finish times."""
        start = "2026-05-06 08:00:00"
        finish = "2026-05-06 09:30:45"
        elapsed = compute_elapsed(finish, start)
        assert elapsed == "1:30:45"
    
    def test_compute_elapsed_over_hour(self):
        """Test elapsed time over multiple hours."""
        start = "2026-05-06 08:00:00"
        finish = "2026-05-06 11:15:30"
        elapsed = compute_elapsed(finish, start)
        assert elapsed == "3:15:30"
    
    def test_compute_elapsed_under_hour(self):
        """Test elapsed time under one hour."""
        start = "2026-05-06 08:00:00"
        finish = "2026-05-06 08:45:20"
        elapsed = compute_elapsed(finish, start)
        assert elapsed == "0:45:20"
    
    def test_compute_elapsed_missing_finish(self):
        """Test with missing finish time."""
        start = "2026-05-06 08:00:00"
        elapsed = compute_elapsed(None, start)
        assert elapsed is None
    
    def test_compute_elapsed_missing_start(self):
        """Test with missing start time."""
        finish = "2026-05-06 09:30:00"
        elapsed = compute_elapsed(finish, None)
        assert elapsed is None
    
    def test_compute_elapsed_negative_time(self):
        """Test with finish time before start time."""
        start = "2026-05-06 10:00:00"
        finish = "2026-05-06 09:00:00"
        elapsed = compute_elapsed(finish, start)
        assert elapsed is None
    
    def test_compute_elapsed_invalid_format(self):
        """Test with invalid datetime format."""
        start = "invalid-date"
        finish = "2026-05-06 09:00:00"
        elapsed = compute_elapsed(finish, start)
        assert elapsed is None
    
    def test_compute_elapsed_with_t_separator(self):
        """Test with ISO format using T separator."""
        start = "2026-05-06T08:00:00"
        finish = "2026-05-06T09:30:00"
        elapsed = compute_elapsed(finish, start)
        assert elapsed == "1:30:00"


class TestNormalizeRunnerRow:
    """Tests for the normalize_runner_row function."""
    
    def test_normalize_french_columns(self):
        """Test normalizing French column names."""
        row = {
            'dossard': '101',
            'nom': 'Dupont',
            'prenom': 'Jean',
            'sexe': 'M',
            'distance': 'Trail 10K',
            'dossier': 'COMPLET'
        }
        normalized = normalize_runner_row(row)
        
        assert normalized['bib_number'] == '101'
        assert normalized['name'] == 'Jean Dupont'
        assert normalized['gender'] == 'M'
        assert normalized['race'] == 'Trail 10K'
        assert normalized['dossier_complete'] == 1
    
    def test_normalize_english_columns(self):
        """Test normalizing English column names."""
        row = {
            'bib_number': '202',
            'name': 'John Doe',
            'gender': 'M',
            'race': '5K Fun Run',
            'dossier_complete': 1
        }
        normalized = normalize_runner_row(row)
        
        assert normalized['bib_number'] == '202'
        assert normalized['name'] == 'John Doe'
        assert normalized['gender'] == 'M'
        assert normalized['race'] == '5K Fun Run'
        assert normalized['dossier_complete'] == 1
    
    def test_normalize_dossier_complete_variations(self):
        """Test different values for dossier complete status."""
        test_cases = [
            ('COMPLET', 1),
            ('COMPLETE', 1),
            ('1', 1),
            ('YES', 1),
            ('OUI', 1),
            ('INCOMPLET', 0),
            ('0', 0),
            ('NO', 0),
            ('', 0),
        ]
        
        for dossier_val, expected in test_cases:
            row = {'dossier': dossier_val}
            normalized = normalize_runner_row(row)
            assert normalized['dossier_complete'] == expected, \
                f"Failed for dossier value: {dossier_val}"
    
    def test_normalize_missing_dossier(self):
        """Test normalization when dossier field is missing."""
        row = {'bib_number': '303'}
        normalized = normalize_runner_row(row)
        assert normalized['dossier_complete'] == 0
    
    def test_normalize_with_age(self):
        """Test normalization with age field."""
        row = {'bib_number': '404', 'age': '30'}
        normalized = normalize_runner_row(row)
        assert normalized['age'] == '30'
    
    def test_normalize_empty_name_parts(self):
        """Test normalization with empty or None name parts."""
        row = {'nom': '', 'prenom': 'Jean'}
        normalized = normalize_runner_row(row)
        assert normalized['name'] == 'Jean'
        
        row = {'nom': 'Dupont', 'prenom': ''}
        normalized = normalize_runner_row(row)
        assert normalized['name'] == 'Dupont'
        
        row = {'nom': None, 'prenom': None}
        normalized = normalize_runner_row(row)
        assert normalized['name'] == ''


class TestParseUpload:
    """Tests for the parse_upload function."""
    
    def test_parse_csv_file(self):
        """Test parsing a CSV file."""
        csv_content = "bib_number,name,age\n101,John Doe,30\n102,Jane Smith,25\n"
        file_storage = MockFileStorage(csv_content.encode(), 'runners.csv')
        
        rows = parse_upload(file_storage)
        
        assert len(rows) == 2
        assert rows[0]['bib_number'] == '101'
        assert rows[0]['name'] == 'John Doe'
        assert rows[0]['age'] == '30'
        assert rows[1]['bib_number'] == '102'
    
    def test_parse_csv_with_utf8_bom(self):
        """Test parsing CSV with UTF-8 BOM."""
        # The csv reader with utf-8-sig encoding handles the BOM correctly
        csv_content = "bib_number,name\n101,Test Runner\n"
        # Encode with BOM
        content_with_bom = csv_content.encode('utf-8-sig')
        file_storage = MockFileStorage(content_with_bom, 'runners.csv')
        
        rows = parse_upload(file_storage)
        
        assert len(rows) == 1
        # The BOM should be properly handled by utf-8-sig encoding
        assert 'bib_number' in rows[0]
    
    def test_parse_xlsx_file(self):
        """Test parsing an Excel XLSX file."""
        # Create a simple Excel file in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['bib_number', 'name', 'age'])
        ws.append([201, 'Runner One', 28])
        ws.append([202, 'Runner Two', 32])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        file_storage = MockFileStorage(output.getvalue(), 'runners.xlsx')
        
        rows = parse_upload(file_storage)
        
        assert len(rows) == 2
        assert rows[0]['bib_number'] == '201'
        assert rows[0]['name'] == 'Runner One'
        assert rows[0]['age'] == '28'
    
    def test_parse_xlsx_with_none_values(self):
        """Test parsing Excel file with None/empty values."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['bib_number', 'name', 'age'])
        ws.append([301, 'Test Runner', None])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        file_storage = MockFileStorage(output.getvalue(), 'runners.xlsx')
        
        rows = parse_upload(file_storage)
        
        assert len(rows) == 1
        assert rows[0]['age'] == ''  # None should be converted to empty string
    
    def test_parse_empty_xlsx(self):
        """Test parsing an empty Excel file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        file_storage = MockFileStorage(output.getvalue(), 'runners.xlsx')
        
        rows = parse_upload(file_storage)
        
        assert rows == []
    
    def test_parse_unsupported_file_type(self):
        """Test parsing an unsupported file type."""
        file_storage = MockFileStorage(b'some content', 'file.txt')
        
        with pytest.raises(ValueError) as exc_info:
            parse_upload(file_storage)
        
        assert 'Unsupported file type' in str(exc_info.value)
    
    def test_parse_csv_case_insensitive_headers(self):
        """Test that CSV headers are case-insensitive."""
        csv_content = "BIB_NUMBER,NAME,AGE\n101,Test,30\n"
        file_storage = MockFileStorage(csv_content.encode(), 'runners.csv')
        
        rows = parse_upload(file_storage)
        
        assert 'bib_number' in rows[0]  # Should be lowercase
        assert 'name' in rows[0]
        assert 'age' in rows[0]
    
    def test_parse_xlsx_case_insensitive_headers(self):
        """Test that Excel headers are case-insensitive."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['BIB_NUMBER', 'NAME'])
        ws.append([401, 'Test'])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        file_storage = MockFileStorage(output.getvalue(), 'test.xlsx')
        
        rows = parse_upload(file_storage)
        
        assert 'bib_number' in rows[0]  # Should be lowercase


class MockFileStorage:
    """Mock file storage object for testing."""
    
    def __init__(self, content, filename):
        self.filename = filename
        self._stream = io.BytesIO(content)
    
    @property
    def stream(self):
        """Return the stream for access."""
        return self._stream
    
    @stream.setter
    def stream(self, value):
        """Allow setting stream (needed for some operations)."""
        self._stream = value
    
    def __getattr__(self, name):
        """Delegate all other attributes to the underlying stream."""
        return getattr(self._stream, name)


class TestIntegrationUtilityFunctions:
    """Integration tests for utility functions used together."""
    
    def test_full_import_workflow_french(self):
        """Test complete workflow: parse CSV and normalize French data."""
        csv_content = "DOSSARD,NOM,PRENOM,SEXE,DISTANCE,DOSSIER\n101,Dupont,Jean,M,10K,COMPLET\n"
        file_storage = MockFileStorage(csv_content.encode(), 'runners.csv')
        
        rows = parse_upload(file_storage)
        normalized = normalize_runner_row(rows[0])
        
        assert normalized['bib_number'] == '101'
        assert normalized['name'] == 'Jean Dupont'
        assert normalized['gender'] == 'M'
        assert normalized['race'] == '10K'
        assert normalized['dossier_complete'] == 1
    
    def test_elapsed_time_in_ranking_context(self):
        """Test elapsed time calculation as it would be used in ranking."""
        race_start = "2026-05-06 08:00:00"
        finish_times = [
            "2026-05-06 09:30:00",
            "2026-05-06 09:45:00",
            "2026-05-06 08:30:00",
        ]
        
        elapsed_times = [
            compute_elapsed(ft, race_start) for ft in finish_times
        ]
        
        assert elapsed_times[0] == "1:30:00"
        assert elapsed_times[1] == "1:45:00"
        assert elapsed_times[2] == "0:30:00"
        
        # Verify sorting would work correctly
        assert elapsed_times[2] < elapsed_times[0] < elapsed_times[1]
